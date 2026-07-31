import io
import re
import openpyxl
from fastapi import HTTPException, UploadFile
from config import get_settings
from deps import get_supabase_admin

settings = get_settings()

# Maps registrar sheet-tab names (as used in official enrollment/graduating lists)
# to the exact course values used elsewhere in the app (dropdowns, priority logic, etc).
# Keys are normalized: uppercase, no parenthetical suffixes (e.g. "(2)"), no extra whitespace.
SHEET_COURSE_MAP = {
    "BSIT": "Bachelor of Science in Information Technology",
    "BSHM": "Bachelor of Science in Hospitality Management",
    "BSCRIM": "Bachelor of Science in Criminology",
    "BEED": "Bachelor of Elementary Education",
    "BSTM": "Bachelor of Science in Tourism Management",
    "BS-TM": "Bachelor of Science in Tourism Management",
    "BSBA": "Bachelor of Science in Business Administration",
    "BSBA-FM": "Bachelor of Science in Business Administration",
    "BSBAFM": "Bachelor of Science in Business Administration",
    "BSA": "Bachelor of Science in Accountancy",
    # BSEd majors — each sheet is a major under Secondary Education, not its own course
    "ENGLISH": "Bachelor of Secondary Education",
    "FILIPINO": "Bachelor of Secondary Education",
    "MATH": "Bachelor of Secondary Education",
    "SCIENCE": "Bachelor of Secondary Education",
    "SOCIAL STUDIES": "Bachelor of Secondary Education",
    "PSYCHOLOGY": "Bachelor of Secondary Education",
}

# Header cell labels we look for while scanning down a sheet for its header row.
# Each maps to the field name we use internally. "surname" covers the registrar's
# own terminology; "last name" covers the generic format some sheets may use.
HEADER_FIELD_ALIASES = {
    "student id": "student_id",
    "surname": "last_name",
    "last name": "last_name",
    "first name": "first_name",
    "priority class": "priority_class",
}


def _normalize_sheet_name(name: str) -> str:
    # Strip parenthetical suffixes like "(2)" and collapse whitespace
    cleaned = re.sub(r"\(.*?\)", "", name)
    return cleaned.strip().upper()


def _normalize_cell(value) -> str:
    return str(value).strip().lower() if value is not None else ""


def _merged_span(sheet, row_idx_0based, col_idx_0based):
    """Registrar sheets often center a header label over several merged columns
    (e.g. a 'STUDENT ID' header spanning cols C:E), while the actual data value
    for that row lands in just one column somewhere within that span rather than
    the merge's anchor column. Returns (min_col, max_col), 0-indexed inclusive,
    for the merged range containing this cell, or (col, col) if it isn't merged."""
    excel_row = row_idx_0based + 1
    excel_col = col_idx_0based + 1
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= excel_row <= merged_range.max_row and \
           merged_range.min_col <= excel_col <= merged_range.max_col:
            return merged_range.min_col - 1, merged_range.max_col - 1
    return col_idx_0based, col_idx_0based


def _find_header_row(sheet, rows):
    """Scan down a sheet's rows for the header row (contains a 'student id' cell).
    Returns (header_row_index, {field_name: (min_col, max_col)}) or (None, {}) if not found."""
    for i, row in enumerate(rows):
        normalized_cells = [_normalize_cell(c) for c in row]
        if "student id" not in normalized_cells:
            continue
        col_map = {}
        for col_idx, cell in enumerate(normalized_cells):
            if not cell:
                continue
            for label, field in HEADER_FIELD_ALIASES.items():
                if field in col_map:
                    continue
                if label in cell or cell in label:
                    col_map[field] = _merged_span(sheet, i, col_idx)
        # A usable header row needs at minimum student_id, first_name, last_name
        if {"student_id", "first_name", "last_name"} <= col_map.keys():
            return i, col_map
    return None, {}


def _read_span_value(row, span):
    """Given a (min_col, max_col) span, return the first non-empty cell value in it."""
    min_col, max_col = span
    for col in range(min_col, max_col + 1):
        if col < len(row) and row[col] is not None and str(row[col]).strip() != "":
            return row[col]
    return None


async def upload_student_records(file: UploadFile, default_priority: str = "regular") -> dict:
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    contents = await file.read()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

    admin = get_supabase_admin()
    records_to_upsert = []
    skipped_sheets = []
    processed_sheets = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            skipped_sheets.append(f"{sheet_name} (empty)")
            continue

        course = SHEET_COURSE_MAP.get(_normalize_sheet_name(sheet_name))
        if not course:
            skipped_sheets.append(f"{sheet_name} (unrecognized course/tab name)")
            continue

        header_idx, col_map = _find_header_row(sheet, rows)
        if header_idx is None:
            skipped_sheets.append(f"{sheet_name} (no header row found)")
            continue

        sheet_count = 0
        for row in rows[header_idx + 1:]:
            if not any(row):  # skip blank rows
                continue

            raw_student_id = _read_span_value(row, col_map["student_id"])
            student_id = str(raw_student_id).strip() if raw_student_id is not None else ""
            if not student_id or student_id == 'None':
                continue

            priority = default_priority
            if "priority_class" in col_map:
                priority_val = _read_span_value(row, col_map["priority_class"])
                if priority_val:
                    priority = str(priority_val).strip().lower()

            first_name_val = _read_span_value(row, col_map["first_name"])
            last_name_val = _read_span_value(row, col_map["last_name"])

            # Guard against stray annotation rows (e.g. a leftover note sitting alone
            # in the student-id column) that aren't real student records.
            if not first_name_val and not last_name_val:
                continue

            records_to_upsert.append({
                "student_id": student_id,
                "first_name": str(first_name_val).strip() if first_name_val else "",
                "last_name": str(last_name_val).strip() if last_name_val else "",
                "course": course,
                "priority_class": priority
            })
            sheet_count += 1

        if sheet_count:
            processed_sheets.append(f"{sheet_name} ({sheet_count})")
        else:
            skipped_sheets.append(f"{sheet_name} (no data rows)")

    if not records_to_upsert:
        detail = "No valid records found to import."
        if skipped_sheets:
            detail += " Skipped sheets: " + ", ".join(skipped_sheets)
        raise HTTPException(status_code=400, detail=detail)

    # Supabase limits bulk inserts. Let's do batches of 1000
    batch_size = 1000
    inserted = 0
    try:
        for i in range(0, len(records_to_upsert), batch_size):
            batch = records_to_upsert[i:i+batch_size]
            admin.table("school_students").upsert(batch).execute()
            inserted += len(batch)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error during insert: {str(e)}")

    message = f"Successfully imported {inserted} student records from {len(processed_sheets)} sheet(s)."
    if skipped_sheets:
        message += f" Skipped {len(skipped_sheets)} sheet(s): " + ", ".join(skipped_sheets)

    return {"message": message}


async def add_student_record(student_id: str, first_name: str, last_name: str, course: str, priority_class: str) -> dict:
    admin = get_supabase_admin()
    try:
        admin.table("school_students").upsert({
            "student_id": student_id,
            "first_name": first_name,
            "last_name": last_name,
            "course": course,
            "priority_class": priority_class
        }).execute()
        return {"message": "Record added successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to add record: {str(e)}")


async def get_student_records() -> dict:
    admin = get_supabase_admin()
    try:
        res = admin.table("school_students").select("*").order("created_at", desc=True).limit(500).execute()
        return {"records": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch records: {str(e)}")

async def delete_student_record(student_id: str) -> dict:
    admin = get_supabase_admin()
    try:
        admin.table("school_students").delete().eq("student_id", student_id).execute()
        return {"message": "Record deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete record: {str(e)}")

async def bulk_delete_student_records(student_ids: list[str]) -> dict:
    if not student_ids:
        raise HTTPException(status_code=400, detail="No student IDs provided for deletion.")
    
    admin = get_supabase_admin()
    try:
        admin.table("school_students").delete().in_("student_id", student_ids).execute()
        return {"message": f"Successfully deleted {len(student_ids)} record(s)"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to bulk delete records: {str(e)}")

async def update_student_record(student_id: str, first_name: str, last_name: str, course: str, priority_class: str) -> dict:
    admin = get_supabase_admin()
    try:
        admin.table("school_students").update({
            "first_name": first_name,
            "last_name": last_name,
            "course": course,
            "priority_class": priority_class
        }).eq("student_id", student_id).execute()
        
        # Also try to update their actual user profile if they have one
        try:
            admin.table("users").update({
                "first_name": first_name,
                "last_name": last_name,
                "course": course,
                "priority_class": priority_class
            }).eq("student_id", student_id).execute()
        except Exception:
            pass # Ignore if they don't have a user account yet
            
        return {"message": "Record updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update record: {str(e)}")